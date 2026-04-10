import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from io import BytesIO

# ---------------------- 页面配置 ----------------------
st.set_page_config(
    page_title="包装清单生成工具",
    page_icon="📦",
    layout="centered"
)

# ---------------------- 标题和输入框 ----------------------
st.title("📦 包装清单自动生成工具")
st.divider()

# 输入区域（简洁美观）
col1, col2, col3 = st.columns(3)
with col1:
    q1 = st.number_input(
        "485213550SXJ\n(Inner tie rod) 总数量",
        min_value=0.0,
        step=1.0,
        value=4000.0,
        format="%.0f"
    )
with col2:
    q2 = st.number_input(
        "485203550SXJ\n(Outer tie rod) 总数量",
        min_value=0.0,
        step=1.0,
        value=2000.0,
        format="%.0f"
    )
with col3:
    q3 = st.number_input(
        "486403550SXJ\n(Outer tie rod) 总数量",
        min_value=0.0,
        step=1.0,
        value=2000.0,
        format="%.0f"
    )

# ---------------------- 生成Excel核心函数 ----------------------
def generate_excel(q1, q2, q3):
    wb = Workbook()
    ws = wb.active
    
    # 样式定义
    thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    wrap_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    font11b = Font(name="MS PGothic", size=11, bold=True)
    font12b = Font(name="MS PGothic", size=12, bold=True)
    font14b = Font(name="MS PGothic", size=14, bold=True)
    font16b = Font(name="MS PGothic", size=16, bold=True)
    font20b = Font(name="MS PGothic", size=20, bold=True)

    # 基础信息
    ws['A1'] = "THK Rhythm(Changzhou) Co.,Ltd."
    ws['E2'] = "Detail Packing List"
    ws['M3'] = f"DATE: {datetime.today().strftime('%Y/%m/%d')}"
    ws['A5'] = "Invoice No:"
    ws['C5'] = "TRCC-26XXX"
    ws['A6'] = "As attached"
    
    ws['A1'].font = font14b
    ws['E2'].font = font20b
    ws['M3'].font = font12b
    ws['C5'].font = font12b
    ws['A6'].font = font16b
    ws['A1'].alignment = ws['A5'].alignment = ws['A6'].alignment = left_align

    # 表头设置
    headers = [
        'P.T.R No', 'PALLET NO.', 'CASE NO.（CN NO.）', 'PART NO.', 'PART NAME',
        "Q'TY(PCS)/CTN", "TOTAL Q'TY(PCS)", "WEIGHT(KG)/CARTON", "",
        "TOTAL WEIGHT(KG)", "", "CBM(m³)( L*W*H )", "REMARKS"
    ]
    for col, v in enumerate(headers, 1):
        c = ws.cell(row=7, column=col, value=v)
        c.font = font12b
        c.alignment = wrap_center
        c.border = thin_border

    # 合并单元格
    for col in [1,2,3,4,5,6]:
        ws.merge_cells(start_row=7, start_column=col, end_row=8, end_column=col)
    ws.merge_cells('G7:G8')
    ws.merge_cells('H7:I7')
    ws['H8'] = 'NET'
    ws['I8'] = 'GROSS'
    ws.merge_cells('J7:K7')
    ws['J8'] = 'NET'
    ws['K8'] = 'GROSS'
    ws.merge_cells('L7:L8')
    ws.merge_cells('M7:M8')
    
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 20

    # 表头样式补充
    for r in [7,8]:
        for col in range(1,14):
            c = ws.cell(row=r, column=col)
            c.border = thin_border
            c.font = font12b
            c.alignment = wrap_center

    # 物料配置
    materials = [
        {"pn":"485213550SXJ","name":"Inner tie rod","total":q1,"pallet":800,"ctn":20,"net":12.4,"gross":13,"t_net":496,"t_gross":535},
        {"pn":"485203550SXJ","name":"Outer tie rod","total":q2,"pallet":400,"ctn":10,"net":5.6,"gross":6.12,"t_net":224,"t_gross":259.8},
        {"pn":"486403550SXJ","name":"Outer tie rod","total":q3,"pallet":400,"ctn":10,"net":5.6,"gross":6.12,"t_net":224,"t_gross":259.8},
    ]

    # 填充数据（含余数逻辑）
    row = 9
    pallet_no = 1
    data_rows = []
    cbm_text = "SIZE:1.2*0.8*0.83m³"

    def write_row(r, pn, name, qty, ctn_qty, net, gross, t_net, t_gross, p_num):
        data_rows.append(r)
        ws.cell(r,1, ''), ws.cell(r,2, p_num), ws.cell(r,3, f'1-{int(qty/ctn_qty)}')
        ws.cell(r,4, pn), ws.cell(r,5, name), ws.cell(r,6, ctn_qty), ws.cell(r,7, qty)
        ws.cell(r,8, net), ws.cell(r,9, gross), ws.cell(r,10, t_net), ws.cell(r,11, t_gross)
        ws.cell(r,12, cbm_text), ws.cell(r,13, f"PALLET NO.{p_num}")
        
        for col in range(1,14):
            c = ws.cell(r, col)
            c.font = font11b
            c.alignment = center
            c.border = thin_border
            if col == 2: c.number_format = '0'
            if col in [6,7,8,9,10,11,12]: c.number_format = '0.00'

    for m in materials:
        total = m['total']
        per_pallet = m['pallet']
        full = int(total // per_pallet)
        rem = total % per_pallet

        # 整托行
        for _ in range(full):
            write_row(row, m['pn'], m['name'], per_pallet, m['ctn'], m['net'], m['gross'], m['t_net'], m['t_gross'], pallet_no)
            row +=1
            pallet_no +=1

        # 余数行
        if rem > 0:
            r_net = round(rem/per_pallet * m['t_net'],2)
            r_gross = round(rem/per_pallet * m['t_gross'],2)
            write_row(row, m['pn'], m['name'], rem, m['ctn'], m['net'], m['gross'], r_net, r_gross, pallet_no)
            row +=1
            pallet_no +=1

    # 合计行
    total_row = row
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws.cell(total_row, 1, f"TOTAL {len(data_rows)} PALLET(S)")
    
    ws.cell(total_row,7,f"=SUM(G{data_rows[0]}:G{data_rows[-1]})")
    ws.cell(total_row,10,f"=SUM(J{data_rows[0]}:J{data_rows[-1]})")
    ws.cell(total_row,11,f"=SUM(K{data_rows[0]}:K{data_rows[-1]})")
    ws.cell(total_row,12, len(data_rows)*0.7968)

    # 合计行样式
    for col in range(1,14):
        c = ws.cell(total_row, col)
        c.font = font11b
        c.alignment = center
        c.border = thin_border
        if col ==2: c.number_format='@'
        if col in [7,10,11,12]: c.number_format='0.00'

    # 列宽设置
    widths = [15.13,13.13,15.5,15.63,36.13,17.63,19.88,11.38,11.38,11.38,11.38,23,20.13]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[chr(64+i)].width = w

    # 保存到BytesIO（用于下载）
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ---------------------- 生成按钮和下载 ----------------------
st.divider()
if st.button("✅ 生成并下载Excel文件", type="primary"):
    with st.spinner("正在生成Excel文件..."):
        excel_buf = generate_excel(q1, q2, q3)
        # 提供下载按钮
        st.download_button(
            label="📥 点击下载 Detail(XNF).xlsx",
            data=excel_buf,
            file_name="Detail(XNF).xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("Excel文件生成成功！点击上方按钮下载～")
